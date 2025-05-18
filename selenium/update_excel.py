import openpyxl
def update_excel_data(filepath,searchterm,colname,new_value):
    filepath = "D:\Pragya Sharma\Downloads\python demo.xlsx"
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    Dict={}
    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value==colname:
            Dict["col"]=i
    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j)==searchterm:
                Dict["row"]=i
    sheet.cell(row=Dict["row"],column=Dict["col"]).value=new_value
    book.save(filepath)
update_excel_data(filepath,"Apple","price","999")
